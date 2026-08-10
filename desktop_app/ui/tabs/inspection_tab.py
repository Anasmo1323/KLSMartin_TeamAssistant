import pandas as pd
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                             QLineEdit, QPushButton, QHeaderView, QFileDialog, 
                             QMessageBox, QDialog, QTableWidgetItem, QComboBox)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QIntValidator

from ui.widgets.segmented_edit import SegmentedCodeEdit
from ui.widgets.dynamic_table import DynamicTableWidget
from ui.dialogs.mapping_dialog import MappingDialog
from core.utils import show_loading

class InspectionTab(QWidget):
    def __init__(self, master_tab):
        super().__init__()
        self.master_tab = master_tab
        self.df = None 
        self.extras = []
        self.display_cols = []
        self.current_file_path = None
        self.is_modified = False
        
        self.mapped_header_row = 0
        self.mapped_sheet_name = None
        self.mapped_code_col = None
        
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(8)

        top_layout = QHBoxLayout()
        top_layout.setSpacing(8)
        self.btn_upload = QPushButton("Upload Verification Manifest")
        self.btn_upload.clicked.connect(self.upload_manifest)
        top_layout.addWidget(self.btn_upload)
        
        self.btn_export_results = QPushButton("Save Inspection Results")
        self.btn_export_results.clicked.connect(self.save_inspection_results)
        self.btn_export_results.setEnabled(False)
        top_layout.addWidget(self.btn_export_results)
        layout.addLayout(top_layout)

        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("<b>Global Filter:</b>"))
        self.txt_global_search = QLineEdit()
        self.txt_global_search.returnPressed.connect(self.apply_filter)
        filter_layout.addWidget(self.txt_global_search)
        
        self.filter_combo = QComboBox()
        self.filter_combo.addItem("All Items", None)
        self.filter_combo.addItem("Code Not Found", "not_found")
        self.filter_combo.addItem("No Image", "no_image")
        self.filter_combo.currentIndexChanged.connect(self.apply_filter)
        filter_layout.addWidget(self.filter_combo)
        
        self.btn_search = QPushButton("🔍")
        self.btn_search.setFixedWidth(50)
        self.btn_search.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_search.clicked.connect(self.apply_filter)
        filter_layout.addWidget(self.btn_search)
        
        layout.addLayout(filter_layout)

        # Enhanced Verification Input Row (Includes Quantity Flow)
        verify_layout = QHBoxLayout()
        verify_layout.addWidget(QLabel("<b>Code Found:</b>"))
        self.code_verify_input = SegmentedCodeEdit()
        
        # When code is entered, shift focus to the Quantity box
        self.code_verify_input.returnPressed.connect(self.focus_qty_box)
        self.code_verify_input.textChanged.connect(self.apply_filter)
        verify_layout.addWidget(self.code_verify_input)

        verify_layout.addWidget(QLabel("<b>Qty:</b>"))
        self.qty_input = QLineEdit("1")
        self.qty_input.setValidator(QIntValidator(1, 9999))
        self.qty_input.setFixedWidth(60)
        self.qty_input.setAlignment(Qt.AlignmentFlag.AlignCenter)
        # When Enter is pressed on Qty, execute the batch verification
        self.qty_input.returnPressed.connect(self.process_verification)
        verify_layout.addWidget(self.qty_input)

        self.btn_verify = QPushButton("Confirm Batch")
        self.btn_verify.setObjectName("primaryButton")
        self.btn_verify.clicked.connect(self.process_verification)
        verify_layout.addWidget(self.btn_verify)
        layout.addLayout(verify_layout)

        self.table = DynamicTableWidget()
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setDefaultSectionSize(130)
        self.table.itemChanged.connect(self.on_item_changed)
        layout.addWidget(self.table)

    def upload_manifest(self):
        if self.is_modified:
            reply = QMessageBox.question(self, 'Unsaved Changes', 
                "You have unsaved changes. Discard and upload new?", 
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.No: return

        file_path, _ = QFileDialog.getOpenFileName(self, "Open Manifest File", "", "Data Files (*.csv *.xlsx *.xls)")
        if not file_path: return

        dlg = MappingDialog(file_path, ['code', 'quantity'], allow_extras=True, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            with show_loading(self, "Processing Manifest File..."):
                try:
                    if dlg.is_excel:
                        raw_df = pd.read_excel(file_path, sheet_name=dlg.selected_sheet, header=dlg.header_row)
                    else:
                        raw_df = pd.read_csv(file_path, header=dlg.header_row)

                    inverted_map = {v: k for k, v in dlg.mappings.items()}
                    cols_to_keep = list(dlg.mappings.values()) + getattr(dlg, 'extras', [])
                    self.df = raw_df[cols_to_keep].rename(columns=inverted_map)
                    self.extras = getattr(dlg, 'extras', [])
                    self.current_file_path = file_path
                    
                    self.mapped_header_row = dlg.header_row
                    self.mapped_sheet_name = getattr(dlg, 'selected_sheet', None)
                    self.mapped_code_col = dlg.mappings['code']
                    
                    self.df['quantity'] = pd.to_numeric(self.df['quantity'], errors='coerce').fillna(0).astype(int)
                    if 'inspected' not in self.df.columns:
                        self.df['inspected'] = 0 
                    if 'status' not in self.df.columns:
                        self.df['status'] = "PENDING"
                    
                    self.set_modified(False)
                    self.btn_export_results.setEnabled(True)
                    self.populate_table(self.df)
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Failed processing manifest: {e}")

    def populate_table(self, dataframe):
        self.table.blockSignals(True)
        self.table.setRowCount(0)
        if dataframe is None or dataframe.empty:
            self.table.blockSignals(False)
            return

        self.display_cols = ['code', 'quantity'] + self.extras + ['inspected', 'status']
        self.table.setColumnCount(len(self.display_cols))
        self.table.setHorizontalHeaderLabels([c.upper() for c in self.display_cols])

        self.table.setRowCount(len(dataframe))
        for ui_row_idx, (orig_idx, row) in enumerate(dataframe.iterrows()):
            status = str(row.get('status', 'PENDING'))
            
            for c_idx, col_name in enumerate(self.display_cols):
                val = str(row.get(col_name, ''))
                item = QTableWidgetItem(val)
                item.setData(Qt.ItemDataRole.UserRole, orig_idx)
                
                # Make status strictly read-only, others are double-click editable
                if col_name == 'status':
                    item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
                else:
                    item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEditable | Qt.ItemFlag.ItemIsEnabled)
                
                if status == "CLEARED":
                    item.setBackground(QColor("#D4EDDA")) 
                    item.setForeground(QColor("#155724"))
                
                self.table.setItem(ui_row_idx, c_idx, item)
        self.table.blockSignals(False)
        
        if self.table.rowCount() > 0:
            self.table.selectRow(0)
            self.table.setCurrentCell(0, 0)

    def on_item_changed(self, item):
        orig_idx = item.data(Qt.ItemDataRole.UserRole)
        if orig_idx is not None and self.df is not None:
            col_name = self.display_cols[item.column()]
            val = item.text()
            
            # Prevent Pandas crash: Cast to int if editing a numeric column
            if col_name in ['quantity', 'inspected']:
                try:
                    val = int(val)
                except ValueError:
                    val = 0
                    
            self.df.at[orig_idx, col_name] = val
            self.set_modified(True)

            # Auto-update status if manual qty/inspected edits occurred
            if col_name in ['quantity', 'inspected']:
                try:
                    q = int(self.df.at[orig_idx, 'quantity'])
                    i = int(self.df.at[orig_idx, 'inspected'])
                    new_status = "CLEARED" if i >= q else "PENDING"
                    if self.df.at[orig_idx, 'status'] != new_status:
                        self.df.at[orig_idx, 'status'] = new_status
                        self.apply_filter() # Refresh table colors
                except: pass

    def focus_qty_box(self):
        """Moves cursor to quantity box and highlights text for quick typing"""
        self.qty_input.setFocus()
        self.qty_input.selectAll()

    def process_verification(self):
        if self.df is None or self.df.empty: return

        target_code = self.code_verify_input.get_code().strip().replace('-', '').lower()
        if not target_code: return

        mask = self.df['code'].astype(str).str.replace('-', '').str.strip().str.lower() == target_code
        indices = self.df[mask].index

        if len(indices) == 0:
            QMessageBox.critical(self, "Invalid Match", "Code not found in client manifest.")
            self.code_verify_input.blockSignals(True)
            for edit in self.code_verify_input.edits: edit.blockSignals(True)
            self.code_verify_input.clear()
            for edit in self.code_verify_input.edits: edit.blockSignals(False)
            self.code_verify_input.blockSignals(False)
            self.code_verify_input.edits[0].setFocus()
            return

        try:
            qty_to_add = int(self.qty_input.text())
        except:
            qty_to_add = 1

        total_capacity = 0
        available_rows = []
        for idx in indices:
            c_inspected = int(self.df.at[idx, 'inspected'])
            t_max = int(self.df.at[idx, 'quantity'])
            cap = max(0, t_max - c_inspected)
            total_capacity += cap
            if cap > 0:
                available_rows.append((idx, cap))

        if qty_to_add > total_capacity:
            QMessageBox.warning(self, "Limit Exceeded", f"Cannot add {qty_to_add}. Only {total_capacity} more needed across all duplicates for this code.")
            self.code_verify_input.blockSignals(True)
            for edit in self.code_verify_input.edits: edit.blockSignals(True)
            self.code_verify_input.clear()
            for edit in self.code_verify_input.edits: edit.blockSignals(False)
            self.code_verify_input.blockSignals(False)
            self.code_verify_input.edits[0].setFocus()
            return

        remaining = qty_to_add
        for idx, cap in available_rows:
            if remaining <= 0:
                break
            add_here = min(cap, remaining)
            c_inspected = int(self.df.at[idx, 'inspected'])
            new_inspected = c_inspected + add_here
            self.df.at[idx, 'inspected'] = new_inspected
            if new_inspected == int(self.df.at[idx, 'quantity']):
                self.df.at[idx, 'status'] = "CLEARED"
            remaining -= add_here
            
            # Manually update UI so we don't need a full redraw
            for ui_row in range(self.table.rowCount()):
                item = self.table.item(ui_row, 0)
                if item and item.data(Qt.ItemDataRole.UserRole) == idx:
                    inspected_col = self.display_cols.index('inspected')
                    status_col = self.display_cols.index('status')
                    self.table.item(ui_row, inspected_col).setText(str(self.df.at[idx, 'inspected']))
                    status = self.df.at[idx, 'status']
                    self.table.item(ui_row, status_col).setText(status)
                    if status == "CLEARED":
                        from PyQt6.QtGui import QColor
                        for c in range(self.table.columnCount()):
                            self.table.item(ui_row, c).setBackground(QColor("#D4EDDA"))
                            self.table.item(ui_row, c).setForeground(QColor("#155724"))
                    break

        self.set_modified(True)
        
        # Workflow Reset (with signals blocked so the table doesn't reset)
        self.code_verify_input.blockSignals(True)
        for edit in self.code_verify_input.edits:
            edit.blockSignals(True)
        self.code_verify_input.clear()
        for edit in self.code_verify_input.edits:
            edit.blockSignals(False)
        self.code_verify_input.blockSignals(False)
        self.qty_input.setText("1")
        self.code_verify_input.edits[0].setFocus()

    def set_modified(self, state):
        self.is_modified = state
        indicator = " *(Unsaved)*" if state else ""
        self.btn_export_results.setText(f"Save Inspection Results{indicator}")

    def save_inspection_results(self):
        if not self.current_file_path: return
        
        try:
            if self.current_file_path.endswith('.csv'):
                export_df = self.df.copy()
                # Restore original column names for CSV export
                original_col_names = {v: k for k, v in self.mapped_dict.items()} if hasattr(self, 'mapped_dict') else {}
                # Add Missing Qty
                export_df['Missing Qty'] = export_df.apply(
                    lambda r: max(0, int(r['quantity']) - int(r['inspected'])) if r['status'] != 'CLEARED' else 0, axis=1)
                export_df.to_csv(self.current_file_path, index=False)
                self.set_modified(False)
                QMessageBox.information(self, "Saved", "Results appended to CSV file.")
            else:
                import openpyxl
                from openpyxl.styles import PatternFill
                
                with show_loading(self, "Applying Highlights to Excel..."):
                    wb = openpyxl.load_workbook(self.current_file_path)
                    ws = wb[self.mapped_sheet_name] if self.mapped_sheet_name else wb.active
                    
                    header_row_excel = self.mapped_header_row + 1
                    code_col_idx = 1
                    max_col = ws.max_column
                    
                    for col in range(1, max_col + 1):
                        cell_val = ws.cell(row=header_row_excel, column=col).value
                        if str(cell_val).strip() == str(self.mapped_code_col).strip():
                            code_col_idx = col
                            break
                            
                    missing_col_idx = max_col + 1
                    ws.cell(row=header_row_excel, column=missing_col_idx, value="Missing Qty")
                    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
                    
                    for df_idx, (orig_idx, row) in enumerate(self.df.iterrows()):
                        excel_row = header_row_excel + 1 + df_idx
                        if row['status'] != "CLEARED":
                            missing = max(0, int(row['quantity']) - int(row['inspected']))
                            ws.cell(row=excel_row, column=missing_col_idx, value=missing)
                            ws.cell(row=excel_row, column=code_col_idx).fill = yellow_fill
                            
                    wb.save(self.current_file_path)
                    self.set_modified(False)
                    QMessageBox.information(self, "Saved", "Highlights and missing quantities saved to original Excel file!")
                    
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save inspection results:\n{e}")

    def apply_filter(self):
        if self.df is None or self.df.empty: return
        g_search = self.txt_global_search.text().strip().lower()
        filtered = self.df.copy()
        
        filter_type = self.filter_combo.currentData()
        if filter_type in ["not_found", "no_image"]:
            master_df = self.master_tab.df
            if master_df is None or master_df.empty:
                self.filter_combo.blockSignals(True)
                self.filter_combo.setCurrentIndex(0)
                self.filter_combo.blockSignals(False)
                QMessageBox.warning(self, "Warning", "Please upload Master CSV first.")
            else:
                clean_codes = filtered['code'].astype(str).str.replace('-', '').str.strip().str.lower()
                master_codes = master_df['code'].astype(str).str.replace('-', '').str.strip().str.lower()
                
                if filter_type == "not_found":
                    mask = ~clean_codes.isin(master_codes)
                    filtered = filtered[mask]
                elif filter_type == "no_image":
                    import os
                    def check_exists(path):
                        if pd.isna(path): return False
                        path = str(path).strip()
                        if path in {"", "No Image", "Download Failed", "HTTP Error", "nan"}: return False
                        return os.path.exists(path)
                        
                    valid_img_mask = master_df['local_image_path'].apply(check_exists)
                    valid_codes_with_img = master_df.loc[valid_img_mask, 'code'].astype(str).str.replace('-', '').str.strip().str.lower()
                    mask = clean_codes.isin(master_codes) & ~clean_codes.isin(valid_codes_with_img)
                    filtered = filtered[mask]
        
        if g_search:
            mask = filtered.astype(str).apply(lambda x: x.str.lower().str.contains(g_search)).any(axis=1)
            filtered = filtered[mask]
            
        code_segs = self.code_verify_input.get_segments()
        active_code = any(s != "" for s in code_segs)
        if active_code:
            code_col = filtered['code'].astype(str).fillna('')
            split_codes = code_col.str.split('-', expand=True)
            for i, seg in enumerate(code_segs):
                if seg:
                    if i < split_codes.shape[1]:
                        filtered = filtered[split_codes[i].str.startswith(seg, na=False)]
                    else:
                        filtered = filtered.iloc[0:0]

        self.populate_table(filtered)
