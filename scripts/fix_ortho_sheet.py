import openpyxl

file_path = 'data/MCC_OfferList_Final.xlsx'
print('Loading workbook...')
wb = openpyxl.load_workbook(file_path)

if 'ORTHO' in wb.sheetnames:
    ws = wb['ORTHO']
    item_counter = 1
    
    # We iterate from row 2 because row 1 is usually the table header (#, CODE, DESCRIPTION, QTY)
    for row in range(2, ws.max_row + 1):
        hash_val = ws.cell(row=row, column=1).value
        code_val = ws.cell(row=row, column=2).value
        
        # Identify a group header: it has a hash_val but no code_val
        if hash_val is not None and (code_val is None or str(code_val).strip() == ''):
            # This is a header row, so we clear the # column
            ws.cell(row=row, column=1).value = None
            # Reset the item counter for the new group
            item_counter = 1
            
        # Identify an item row: it has a code_val but originally no hash_val (we might overwrite it, but we check if it's an item)
        elif code_val is not None and str(code_val).strip() != '':
            # This is an item row, so we set the # column to the counter
            ws.cell(row=row, column=1).value = item_counter
            item_counter += 1

    wb.save(file_path)
    print('ORTHO sheet fixed and saved.')
else:
    print('ORTHO sheet not found!')
