import openpyxl
import re

def renumber_sheets():
    print("Opening existing master file to renumber sheets...")
    master_path = r'C:\Users\Anas Mohamed\PyCharmMiscProject\offer_webapp\master_surgical_sets.xlsx'
    
    wb = openpyxl.load_workbook(master_path)
    
    for idx, ws in enumerate(wb.worksheets, 1):
        # 1. Update the full title in B1
        full_title = ws['B1'].value
        if full_title:
            # Strip existing number prefix (e.g. "4. Aorta Set" -> "Aorta Set")
            clean_full_title = re.sub(r'^\d+[\.\s-]*', '', str(full_title)).strip()
            new_full_title = f"{idx}. {clean_full_title}"
            ws['B1'].value = new_full_title
        else:
            # Fallback if B1 is empty (shouldn't happen with our script, but just in case)
            clean_full_title = re.sub(r'^\d+[\.\s-]*', '', str(ws.title)).strip()
            new_full_title = f"{idx}. {clean_full_title}"
            ws['B1'].value = new_full_title

        # 2. Update the sheet title (tab name)
        # Excel sheet names max 31 characters
        safe_name = re.sub(r'[\\/*?:"<>|]', '', new_full_title)[:31]
        
        # We need to temporarily set it to a dummy name if we might collide with existing numbers
        # But actually openpyxl allows renaming. If we rename "3. Aorta" to "2. Aorta" and there is no 2, it's fine.
        # To completely avoid collision during iteration, we can append a temporary uuid, then rename them all back.
        # But wait, we are iterating in order. A sheet moving from 3 to 2 won't collide with 2 because 2 was already renamed to 1.
        
        # Let's just use a safe renaming loop
        sheet_name = safe_name
        counter = 1
        # Check against OTHER sheets (not itself)
        while any(sheet_name == s.title and s != ws for s in wb.worksheets):
            suffix = f"_{counter}"
            sheet_name = safe_name[:31-len(suffix)] + suffix
            counter += 1
            
        ws.title = sheet_name

    wb.save(master_path)
    print("Sheets renumbered successfully!")

if __name__ == '__main__':
    renumber_sheets()
