import pandas as pd
import openpyxl
import json
import re
from create_master_excel import get_canonical_name, get_items_from_df

def color_tabs():
    print("Parsing sources to determine set origins...")
    dir_path = r'C:\Users\Anas Mohamed\PyCharmMiscProject\SurgicalSets'
    a1_path = dir_path + r'\A1 Sets in details.xlsx'
    a2_path = r'C:\Users\Anas Mohamed\PyCharmMiscProject\offer_webapp\src\data\A2_parsed_sets.json'
    a3_path = dir_path + r'\A3 sets.xlsx'
    
    sets_data = {}
    
    # Process A1
    xls_a1 = pd.ExcelFile(a1_path)
    for sheet in xls_a1.sheet_names:
        cname = get_canonical_name(sheet)
        if cname not in sets_data: sets_data[cname] = set()
        sets_data[cname].add('A1')

    # Process A2
    with open(a2_path, 'r', encoding='utf-8') as f:
        a2_json = json.load(f)
        for s in a2_json:
            cname = get_canonical_name(s['set_name'])
            if cname not in sets_data: sets_data[cname] = set()
            sets_data[cname].add('A2')
            
    # Process A3
    xls_a3 = pd.ExcelFile(a3_path)
    for sheet in xls_a3.sheet_names:
        cname = get_canonical_name(sheet)
        if cname not in sets_data: sets_data[cname] = set()
        sets_data[cname].add('A3')
        
    # Process Demerdash
    demerdash_path = dir_path + r'\Demerdash Order - PO final.xlsx'
    df_d = pd.read_excel(demerdash_path, sheet_name='PO (2)', header=None)
    for _, row in df_d.iterrows():
        art_no = ""
        for val in row.values:
            val_str = str(val).strip()
            if re.search(r'\d{2}-\d{3}-\d{2}-\d{2}', val_str):
                art_no = val_str
                break
                
        desc = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
        if not art_no and desc and desc.lower() != 'nan' and desc.lower() != 'description' and not desc.startswith('Tray Name'):
            cname = get_canonical_name(desc)
            if cname in sets_data or any(kw in desc.lower() for kw in ['set', 'surg', 'instr']):
                if cname not in sets_data: sets_data[cname] = set()
                sets_data[cname].add('Demerdash')
                
    master_path = r'C:\Users\Anas Mohamed\PyCharmMiscProject\offer_webapp\master_surgical_sets.xlsx'
    
    print("Opening existing master file to apply colors...")
    wb = openpyxl.load_workbook(master_path)
    
    for sheet_name in wb.sheetnames:
        # Re-extract the canonical name from the sheet's title in cell B1
        ws = wb[sheet_name]
        full_title = ws['B1'].value
        if not full_title:
            continue
            
        # Remove the leading number e.g. "1. "
        clean_title = re.sub(r'^\d+[\.\s-]*', '', str(full_title))
        cname = get_canonical_name(clean_title)
        
        sources = sets_data.get(cname, set())
        
        has_a = any(s in ['A1', 'A2', 'A3'] for s in sources)
        has_demerdash = 'Demerdash' in sources
        
        # Colors:
        # Blue = '00B0F0'
        # Green = '92D050'
        # Mixed (Yellow or Orange?) -> Let's default to Orange 'FFC000' for merged, unless user specifies
        
        if has_a and has_demerdash:
            ws.sheet_properties.tabColor = 'FFC000' # Orange for mixed
        elif has_a:
            ws.sheet_properties.tabColor = '00B0F0' # Blue
        elif has_demerdash:
            ws.sheet_properties.tabColor = '92D050' # Green
            
    wb.save(master_path)
    print("Tab colors applied successfully!")

if __name__ == '__main__':
    color_tabs()
