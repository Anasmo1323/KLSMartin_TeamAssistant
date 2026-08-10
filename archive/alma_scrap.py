import os
import requests
import pandas as pd
import urllib3
from playwright.sync_api import sync_playwright

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def get_auth_token():
    """Automates the login process to capture the required API token."""
    EMAIL = "albear@technowave-eg.com"
    PASSWORD = "KLSteam@123"
    LOGIN_URL = "https://alma.klsmartin.com/kls-library/quick-converter"
    
    auth_token = None

    print("Initializing browser automation for authentication...")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(ignore_https_errors=True)
        page = context.new_page()

        def capture_token(request):
            nonlocal auth_token
            if "/api/backend/" in request.url:
                headers = request.headers
                if "authorization" in headers:
                    auth_token = headers["authorization"]

        page.on("request", capture_token)

        page.goto(LOGIN_URL)
        
        page.wait_for_selector('input[placeholder="Email *"]')
        page.fill('input[placeholder="Email *"]', EMAIL)
        page.click('button:has-text("Login / Sign up")')

        page.wait_for_selector('input[type="password"]', timeout=15000)
        page.fill('input[type="password"]', PASSWORD)
        page.click('button[type="submit"], button#next')

        print("Waiting for session initialization and token capture...")
        page.wait_for_load_state("networkidle", timeout=30000)
        
        browser.close()
        
    return auth_token

def search_competitor_sku(sku_list, auth_token):
    """Queries the Quick Converter API with the captured token."""
    if not auth_token:
        print("Error: No valid authorization token provided.")
        return None
        
    api_endpoint = "https://alma.klsmartin.com/api/backend/CompetitorProduct/Search"
    
    headers = {
        "Accept": "application/json",
        "Authorization": auth_token,
        "Content-Type": "application/json",
        "Host": "alma.klsmartin.com",
        "Origin": "https://alma.klsmartin.com",
        "Referer": "https://alma.klsmartin.com/kls-library/quick-converter"
    }
    
    payload = {
        "searchTerms": sku_list
    }
    
    print(f"Executing conversion query for {len(sku_list)} items...")
    try:
        response = requests.post(api_endpoint, headers=headers, json=payload, verify=False)
        
        if response.status_code == 200:
            return response.json()
        else:
            print(f"API Request Failed. HTTP Status: {response.status_code}")
            return None
            
    except Exception as e:
        print(f"Request execution error: {e}")
        return None

if __name__ == '__main__':
    import openpyxl
    
    # 1. Automatically fetch the token
    token = get_auth_token()
    
    if token:
        print("Authentication successful.")
        
        file_path = "Set Proposals 11.xlsx"
        print(f"Reading {file_path}...")
        
        try:
            wb = openpyxl.load_workbook(file_path)
            
            # Collect all SKUs to search
            all_skus = set()
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Assuming SKUs are in column C (column index 3)
                for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            sku = str(cell.value).strip()
                            if sku and sku.lower() != 'nan':
                                all_skus.add(sku)
            
            test_skus = list(all_skus)
            print(f"Found {len(test_skus)} unique SKUs to convert.")
            
            sku_to_kls = {}
            if test_skus:
                # Chunk the search if necessary
                chunk_size = 100
                for i in range(0, len(test_skus), chunk_size):
                    chunk = test_skus[i:i+chunk_size]
                    print(f"Processing chunk {i//chunk_size + 1}/{(len(test_skus)-1)//chunk_size + 1}...")
                    result_data = search_competitor_sku(chunk, token)
                    
                    if result_data:
                        for item in result_data:
                            search_term = item.get('searchTerm')
                            matches = item.get('matches', [])
                            if matches:
                                kls_codes = []
                                for match in matches:
                                    kls_products = match.get('matchingKlsMartinProducts') or []
                                    for kls_prod in kls_products:
                                        target_obj = kls_prod.get('product', kls_prod) if isinstance(kls_prod, dict) else kls_prod
                                        if isinstance(target_obj, dict):
                                            kls_sku = target_obj.get('code') or target_obj.get('sku')
                                            if kls_sku:
                                                kls_codes.append(kls_sku)
                                if kls_codes:
                                    # Join multiple codes with a comma, unique ones only
                                    unique_kls = list(dict.fromkeys(kls_codes))
                                    sku_to_kls[search_term] = ", ".join(unique_kls)
                    
            print("Finished API queries. Updating sheets...")
            
            # Update the workbook
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Determine the next empty column
                max_col = ws.max_column
                new_col_idx = max_col + 1
                
                # Add Header
                ws.cell(row=1, column=new_col_idx, value="KLS Martin Code")
                
                for row_idx in range(2, ws.max_row + 1):
                    sku_cell = ws.cell(row=row_idx, column=3)
                    if sku_cell.value and isinstance(sku_cell.value, str):
                        sku = str(sku_cell.value).strip()
                        kls_code = sku_to_kls.get(sku, "")
                        if kls_code:
                            ws.cell(row=row_idx, column=new_col_idx, value=kls_code)
                            
            output_path = "Set Proposals 11_Updated.xlsx"
            wb.save(output_path)
            print(f"Updated Excel file saved to {output_path}")
            
        except Exception as e:
            print(f"An error occurred during processing: {e}")
    else:
        print("Authentication failed. Check credentials or network state.")
